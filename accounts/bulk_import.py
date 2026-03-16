"""
Bulk Import Service for User Onboarding

Handles Excel file parsing and bulk user creation with validation.
"""

import io
from typing import Tuple, List, Dict, Any
from django.db import transaction
from django.contrib.auth.models import User
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import UserProfile
from .utils import generate_username, generate_password


# Required fields per role
ROLE_REQUIRED_FIELDS = {
    'school_staff': ['first_name', 'last_name', 'email', 'phone', 'employee_id'],
    'parent': ['first_name', 'last_name', 'phone'],
    'student': ['first_name', 'last_name', 'admission_no'],
}

# All fields per role (for template generation)
ROLE_ALL_FIELDS = {
    'school_staff': [
        'first_name', 'last_name', 'email', 'phone', 'employee_id',
        'department', 'designation', 'gender', 'date_of_birth'
    ],
    'parent': [
        'first_name', 'last_name', 'email', 'phone',
        'gender', 'address', 'city', 'state', 'postal_code'
    ],
    'student': [
        'first_name', 'last_name', 'email', 'phone', 'admission_no',
        'roll_no', 'gender', 'date_of_birth'
    ],
}

# Sample data for template
ROLE_SAMPLE_DATA = {
    'school_staff': {
        'first_name': 'John',
        'last_name': 'Smith',
        'email': 'john.smith@school.com',
        'phone': '9876543210',
        'employee_id': 'EMP001',
        'department': 'Mathematics',
        'designation': 'Teacher',
        'gender': 'male',
        'date_of_birth': '1985-06-15',
    },
    'parent': {
        'first_name': 'Rahul',
        'last_name': 'Sharma',
        'email': 'rahul.sharma@email.com',
        'phone': '9876543211',
        'gender': 'male',
        'address': '123 Main Street',
        'city': 'Korba',
        'state': 'Chhattisgarh',
        'postal_code': '495677',
    },
    'student': {
        'first_name': 'Amit',
        'last_name': 'Kumar',
        'email': '',
        'phone': '9876543212',
        'admission_no': '2024001',
        'roll_no': '01',
        'gender': 'male',
        'date_of_birth': '2010-03-20',
    },
}


def generate_import_template(role: str) -> bytes:
    """
    Generate an Excel template for bulk import.

    Args:
        role: User role ('school_staff', 'parent', 'student')

    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{role.replace('_', ' ').title()} Import"

    # Get fields for this role
    fields = ROLE_ALL_FIELDS.get(role, [])
    required_fields = ROLE_REQUIRED_FIELDS.get(role, [])
    sample_data = ROLE_SAMPLE_DATA.get(role, {})

    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    required_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    required_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Add headers
    for col, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

        # Mark required fields in row 2 as instructions
        is_required = field in required_fields
        instruction_cell = ws.cell(
            row=2, column=col,
            value="Required" if is_required else "Optional"
        )
        if is_required:
            instruction_cell.fill = required_fill
            instruction_cell.font = required_font
        instruction_cell.alignment = Alignment(horizontal='center')
        instruction_cell.border = border

    # Add sample data row
    for col, field in enumerate(fields, 1):
        cell = ws.cell(row=3, column=col, value=sample_data.get(field, ''))
        cell.border = border

    # Adjust column widths
    for col, field in enumerate(fields, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(15, len(field) + 5)

    # Add notes sheet
    notes_ws = wb.create_sheet("Notes")
    notes = [
        ["Bulk Import Instructions"],
        [""],
        ["1. Fill in user data starting from row 3 (row 2 shows required/optional fields)"],
        ["2. Do not modify column headers in row 1"],
        ["3. Required fields must have values for successful import"],
        ["4. Date format: YYYY-MM-DD (e.g., 2024-01-15)"],
        ["5. Gender values: male, female, other"],
        ["6. Phone numbers should be 10 digits (without country code)"],
        ["7. Maximum 500 rows per import"],
        [""],
        ["Field Descriptions:"],
    ]

    field_descriptions = {
        'first_name': 'First name of the user',
        'last_name': 'Last name of the user',
        'email': 'Email address (unique per user)',
        'phone': 'Phone number (used for SMS credentials)',
        'employee_id': 'Unique employee identifier',
        'department': 'Department name (e.g., Mathematics, Science)',
        'designation': 'Job title (e.g., Teacher, Principal)',
        'admission_no': 'Student admission number (unique)',
        'roll_no': 'Class roll number',
        'gender': 'Gender (male/female/other)',
        'date_of_birth': 'Date of birth (YYYY-MM-DD)',
        'address': 'Street address',
        'city': 'City name',
        'state': 'State name',
        'postal_code': 'PIN/Postal code',
    }

    for field in fields:
        notes.append([f"  - {field}: {field_descriptions.get(field, '')}"])

    for row_num, row_data in enumerate(notes, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = notes_ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)

    notes_ws.column_dimensions['A'].width = 60

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def parse_excel_file(file, role: str) -> Tuple[List[Dict], List[Dict]]:
    """
    Parse Excel file and extract user data.

    Args:
        file: Uploaded file object
        role: User role for validation

    Returns:
        Tuple of (valid_rows, error_rows)
    """
    wb = load_workbook(file)
    ws = wb.active

    fields = ROLE_ALL_FIELDS.get(role, [])
    required_fields = ROLE_REQUIRED_FIELDS.get(role, [])

    # Read header row
    headers = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            headers.append(header.strip().lower())
        else:
            headers.append(None)

    valid_rows = []
    error_rows = []

    # Start from row 3 (skip header and instruction row)
    for row_num in range(3, min(ws.max_row + 1, 503)):  # Max 500 rows
        row_data = {}
        has_data = False

        for col, header in enumerate(headers, 1):
            if header and header in fields:
                value = ws.cell(row=row_num, column=col).value
                if value is not None:
                    has_data = True
                    # Handle date formatting
                    if hasattr(value, 'strftime'):
                        value = value.strftime('%Y-%m-%d')
                    row_data[header] = str(value).strip() if value else ''
                else:
                    row_data[header] = ''

        # Skip empty rows
        if not has_data:
            continue

        # Validate required fields
        missing_fields = []
        for field in required_fields:
            if not row_data.get(field):
                missing_fields.append(field)

        if missing_fields:
            error_rows.append({
                'row': row_num,
                'data': row_data,
                'error': f"Missing required fields: {', '.join(missing_fields)}"
            })
        else:
            row_data['_row'] = row_num
            valid_rows.append(row_data)

    return valid_rows, error_rows


def create_users_from_data(
    valid_rows: List[Dict],
    role: str,
    send_sms: bool = False
) -> Tuple[List[Dict], List[Dict]]:
    """
    Create users from validated row data.

    Uses database transaction to ensure all-or-nothing import.

    Args:
        valid_rows: List of validated row dictionaries
        role: User role
        send_sms: Whether to send SMS notifications

    Returns:
        Tuple of (created_users, errors)
    """
    created_users = []
    errors = []

    # Import SMS service here to avoid circular imports
    from notifications.sms_service import sms_service

    try:
        with transaction.atomic():
            for row_data in valid_rows:
                row_num = row_data.pop('_row', 0)

                try:
                    # Generate credentials
                    if role == 'parent':
                        identifier = row_data.get('phone') or row_data.get('email', '').split('@')[0]
                    elif role == 'student':
                        identifier = row_data.get('admission_no', '')
                    else:  # school_staff
                        identifier = row_data.get('employee_id') or row_data.get('email', '').split('@')[0]

                    username = generate_username(role, identifier)
                    password = generate_password()

                    # Check for duplicate email
                    email = row_data.get('email', '')
                    if email and User.objects.filter(email=email).exists():
                        errors.append({
                            'row': row_num,
                            'data': row_data,
                            'error': f"Email already exists: {email}"
                        })
                        continue

                    # Determine is_staff based on role
                    is_staff = role in ['school_admin', 'school_staff']

                    # Create user
                    user = User.objects.create_user(
                        username=username,
                        email=email,
                        password=password,
                        first_name=row_data.get('first_name', ''),
                        last_name=row_data.get('last_name', ''),
                        is_active=True,
                        is_staff=is_staff,
                    )

                    # Create profile
                    profile = UserProfile.objects.create(
                        user=user,
                        role=role,
                        phone=row_data.get('phone', ''),
                        date_of_birth=row_data.get('date_of_birth') or None,
                        gender=row_data.get('gender', ''),
                        employee_id=row_data.get('employee_id', ''),
                        department=row_data.get('department', ''),
                        designation=row_data.get('designation', ''),
                        admission_no=row_data.get('admission_no', ''),
                        roll_no=row_data.get('roll_no', ''),
                        address=row_data.get('address', ''),
                        city=row_data.get('city', ''),
                        state=row_data.get('state', ''),
                        postal_code=row_data.get('postal_code', ''),
                        must_change_password=True,
                    )

                    user_info = {
                        'row': row_num,
                        'user_id': user.pk,
                        'profile_id': str(profile.id),
                        'username': username,
                        'password': password,  # Include for SMS
                        'name': f"{user.first_name} {user.last_name}".strip(),
                        'phone': row_data.get('phone', ''),
                        'email': email,
                    }

                    created_users.append(user_info)

                except Exception as e:
                    errors.append({
                        'row': row_num,
                        'data': row_data,
                        'error': str(e)
                    })

            # If any errors occurred, rollback everything
            if errors:
                raise Exception("Import failed due to errors")

    except Exception as e:
        # Transaction rolled back, return errors
        if not errors:
            errors.append({'row': 0, 'error': str(e)})
        return [], errors

    # Send SMS notifications if requested
    if send_sms and created_users:
        sms_results = sms_service.send_bulk_credentials([
            {
                'phone': user['phone'],
                'username': user['username'],
                'password': user['password'],
                'name': user['name'],
            }
            for user in created_users
            if user.get('phone')
        ])
        # Attach SMS results to response (optional)

    # Remove passwords from response for security
    for user in created_users:
        # Keep password only if not sending SMS (for manual sharing)
        if send_sms:
            user.pop('password', None)

    return created_users, errors
